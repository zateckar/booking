"""
Dynamic reporting service with configurable columns
Supports both static user data and mapped claims data
"""

import json
import logging
import io
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, text

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from . import models
from .claims_service import ClaimsMappingService
from .logging_config import get_logger

logger = get_logger("dynamic_reports")


class DynamicReportsService:
    """Service for generating reports with configurable columns"""
    
    def __init__(self, db: Session):
        self.db = db
        self.claims_service = ClaimsMappingService(db)
    
    def get_available_columns(self) -> List[Dict[str, Any]]:
        """Get all available columns for reports"""
        # Get configured report columns
        columns = self.db.query(models.ReportColumn).filter(
            models.ReportColumn.is_available == True
        ).order_by(models.ReportColumn.sort_order).all()
        
        # Also discover dynamic columns from existing user profiles
        dynamic_columns = self._discover_dynamic_columns()
        
        # Add built-in calculated columns
        calculated_columns = self._get_built_in_calculated_columns()
        
        # Convert to response format
        result = []
        for column in columns:
            result.append({
                "column_name": column.column_name,
                "display_label": column.display_label,
                "column_type": column.column_type,
                "data_type": column.data_type,
                "is_available": column.is_available,
                "sort_order": column.sort_order
            })
        
        # Add dynamic columns that aren't already configured
        configured_names = {col.column_name for col in columns}
        for dyn_col in dynamic_columns:
            if dyn_col["column_name"] not in configured_names:
                result.append(dyn_col)
        
        # Add calculated columns that aren't already configured
        for calc_col in calculated_columns:
            if calc_col["column_name"] not in configured_names:
                result.append(calc_col)
        
        logger.info(f"Found {len(result)} available report columns")
        return result
    
    def _discover_dynamic_columns(self) -> List[Dict[str, Any]]:
        """Discover available dynamic columns from user profiles and claim mappings"""
        dynamic_columns = []
        field_names = set()
        
        # First, get all configured claim mappings - these should always be available as columns
        claim_mappings = self.db.query(models.OIDCClaimMapping).all()
        for mapping in claim_mappings:
            field_names.add(mapping.mapped_field_name)
        
        # Also get field names from existing user profiles to catch any legacy data
        profiles = self.db.query(models.UserProfile).all()
        for profile in profiles:
            if profile.profile_data:
                try:
                    data = json.loads(profile.profile_data)
                    field_names.update(data.keys())
                except json.JSONDecodeError:
                    continue
        
        # Create column definitions for all discovered fields
        for field_name in field_names:
            # Try to determine data type from claim mappings
            data_type = "string"  # default
            mapping = self.db.query(models.OIDCClaimMapping).filter(
                models.OIDCClaimMapping.mapped_field_name == field_name
            ).first()
            
            if mapping:
                if mapping.mapping_type == "array":
                    data_type = "array"
                elif mapping.mapping_type == "number":
                    data_type = "number"
                elif mapping.mapping_type == "boolean":
                    data_type = "boolean"
                
                display_label = mapping.display_label or field_name.replace("_", " ").title()
            else:
                display_label = field_name.replace("_", " ").title()
            
            dynamic_columns.append({
                "column_name": field_name,
                "display_label": display_label,
                "column_type": "mapped",
                "data_type": data_type,
                "is_available": True,
                "sort_order": 1000  # Put dynamic columns at the end
            })
        
        return dynamic_columns
    
    def _get_built_in_calculated_columns(self) -> List[Dict[str, Any]]:
        """Get built-in calculated columns that are available for reports"""
        return [
            {
                "column_name": "days_with_at_least_one_booking",
                "display_label": "Days with at least one booking",
                "column_type": "calculated",
                "data_type": "number",
                "is_available": True,
                "sort_order": 110  # Place it after other booking-related columns
            }
        ]
    
    def generate_dynamic_report(
        self, 
        selected_columns: List[str], 
        months: int = 2,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict[str, Any]:
        """Generate a report with selected columns"""
        
        # Calculate date range if not provided
        if not start_date or not end_date:
            now = datetime.now(timezone.utc)
            start_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            if months == 0:
                # Special case: Last month only
                # Start of last month
                if start_of_current_month.month == 1:
                    start_date = start_of_current_month.replace(year=start_of_current_month.year - 1, month=12)
                else:
                    start_date = start_of_current_month.replace(month=start_of_current_month.month - 1)
                
                # End of last month (start of current month)
                end_date = start_of_current_month
            else:
                # Original logic: current month + previous months
                start_date = start_of_current_month - timedelta(days=32 * (months - 1))
                start_date = start_date.replace(day=1)
                end_date = now
        
        logger.info(f"Generating dynamic report with {len(selected_columns)} columns for period {start_date} to {end_date}")
        
        # Get bookings in the date range
        bookings = self.db.query(models.Booking).join(models.User).join(models.ParkingSpace).join(models.ParkingLot).filter(
            and_(
                models.Booking.start_time >= start_date,
                models.Booking.start_time <= end_date,
                models.Booking.is_cancelled == False
            )
        ).all()
        
        # Calculate user statistics
        user_stats = self._calculate_user_statistics(bookings)
        
        # Enhance with profile data
        enhanced_stats = self._enhance_with_profile_data(user_stats, selected_columns)
        
        # Calculate summary statistics
        summary = self._calculate_summary_statistics(bookings, enhanced_stats)
        
        # Prepare report data
        report_data = {
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "months": months
            },
            "summary": summary,
            "columns": self._get_column_definitions(selected_columns),
            "data": enhanced_stats,
            "total_records": len(enhanced_stats)
        }
        
        logger.info(f"Generated report with {len(enhanced_stats)} user records")
        return report_data
    
    def _calculate_user_statistics(self, bookings: List[models.Booking]) -> Dict[int, Dict[str, Any]]:
        """Calculate basic user statistics from bookings"""
        user_stats = {}
        
        for booking in bookings:
            user_id = booking.user_id
            if user_id not in user_stats:
                user_stats[user_id] = {
                    'user_id': user_id,
                    'email': booking.user.email,
                    'is_admin': booking.user.is_admin,
                    'total_bookings': 0,
                    'total_hours': 0,
                    'avg_duration': 0,
                    'parking_lots_used': set(),
                    'license_plates_count': set(),
                    'license_plates_list': set(),
                    'booking_dates': set(),
                    'first_booking': booking.start_time,
                    'last_booking': booking.start_time
                }
            
            duration_hours = (booking.end_time - booking.start_time).total_seconds() / 3600
            user_stats[user_id]['total_bookings'] += 1
            user_stats[user_id]['total_hours'] += duration_hours
            user_stats[user_id]['parking_lots_used'].add(booking.space.parking_lot.name)
            
            # Add license plate to both count and list tracking
            if booking.license_plate:  # Only add non-empty license plates
                user_stats[user_id]['license_plates_count'].add(booking.license_plate)
                user_stats[user_id]['license_plates_list'].add(booking.license_plate)
            
            # Track unique booking dates (days with at least one booking)
            booking_date = booking.start_time.date()
            user_stats[user_id]['booking_dates'].add(booking_date)
            
            # Update booking date range
            if booking.start_time < user_stats[user_id]['first_booking']:
                user_stats[user_id]['first_booking'] = booking.start_time
            if booking.start_time > user_stats[user_id]['last_booking']:
                user_stats[user_id]['last_booking'] = booking.start_time
        
        # Calculate averages and convert sets to appropriate formats
        for user_id, stats in user_stats.items():
            if stats['total_bookings'] > 0:
                stats['avg_duration'] = stats['total_hours'] / stats['total_bookings']
            stats['parking_lots_used'] = len(stats['parking_lots_used'])
            
            # Convert license plates to count and sorted list
            stats['license_plates_count'] = len(stats['license_plates_count'])
            stats['license_plates_list'] = sorted(list(stats['license_plates_list']))  # Convert to sorted list for consistency
            
            # Keep backward compatibility with the old 'license_plates' field
            stats['license_plates'] = stats['license_plates_count']
            
            # Calculate days with at least one booking
            stats['days_with_at_least_one_booking'] = len(stats['booking_dates'])
            
            # Remove the temporary booking_dates set as it's not needed in the final output
            del stats['booking_dates']
            
            stats['first_booking'] = stats['first_booking'].isoformat()
            stats['last_booking'] = stats['last_booking'].isoformat()
        
        return user_stats
    
    def _enhance_with_profile_data(self, user_stats: Dict[int, Dict[str, Any]], selected_columns: List[str]) -> List[Dict[str, Any]]:
        """Enhance user statistics with profile data based on selected columns"""
        enhanced_stats = []
        
        # Determine which profile fields are needed
        profile_columns = [col for col in selected_columns if self._is_profile_column(col)]
        
        for user_id, stats in user_stats.items():
            enhanced_record = dict(stats)
            
            # Add profile data if any profile columns are selected
            if profile_columns:
                profile_data = self.claims_service.get_user_profile_data(user_id)
                
                for col in profile_columns:
                    if col in profile_data:
                        enhanced_record[col] = profile_data[col]
                    else:
                        enhanced_record[col] = None
            
            enhanced_stats.append(enhanced_record)
        
        return enhanced_stats
    
    def _is_profile_column(self, column_name: str) -> bool:
        """Check if a column comes from user profile data"""
        static_columns = {
            'user_id', 'email', 'is_admin', 'total_bookings', 'total_hours', 
            'avg_duration', 'parking_lots_used', 'license_plates', 'license_plates_count', 
            'license_plates_list', 'first_booking', 'last_booking', 'days_with_at_least_one_booking'
        }
        return column_name not in static_columns
    
    def _get_column_definitions(self, selected_columns: List[str]) -> List[Dict[str, Any]]:
        """Get column definitions for selected columns"""
        available_columns = self.get_available_columns()
        column_map = {col["column_name"]: col for col in available_columns}
        
        definitions = []
        for col_name in selected_columns:
            if col_name in column_map:
                definitions.append(column_map[col_name])
            else:
                # Create default definition for unknown columns
                definitions.append({
                    "column_name": col_name,
                    "display_label": col_name.replace("_", " ").title(),
                    "column_type": "unknown",
                    "data_type": "string",
                    "is_available": True,
                    "sort_order": 9999
                })
        
        return definitions
    
    def _calculate_summary_statistics(self, bookings: List[models.Booking], user_stats: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate summary statistics for the report"""
        total_bookings = len(bookings)
        unique_users = len(user_stats)
        total_hours = sum(stat['total_hours'] for stat in user_stats)
        avg_booking_duration = total_hours / total_bookings if total_bookings > 0 else 0
        
        return {
            'total_bookings': total_bookings,
            'unique_users': unique_users,
            'total_hours': round(total_hours, 2),
            'avg_booking_duration': round(avg_booking_duration, 2)
        }
    
    def create_report_column(self, column_data: Dict[str, Any]) -> models.ReportColumn:
        """Create a new report column configuration"""
        column = models.ReportColumn(**column_data)
        self.db.add(column)
        self.db.commit()
        self.db.refresh(column)
        
        logger.info(f"Created report column: {column.column_name}")
        return column
    
    def update_report_column(self, column_id: int, update_data: Dict[str, Any]) -> models.ReportColumn:
        """Update a report column configuration"""
        column = self.db.query(models.ReportColumn).filter(models.ReportColumn.id == column_id).first()
        if not column:
            raise ValueError(f"Report column with ID {column_id} not found")
        
        for field, value in update_data.items():
            if hasattr(column, field):
                setattr(column, field, value)
        
        self.db.commit()
        self.db.refresh(column)
        
        logger.info(f"Updated report column {column_id}: {column.column_name}")
        return column
    
    def delete_report_column(self, column_id: int):
        """Delete a report column configuration"""
        column = self.db.query(models.ReportColumn).filter(models.ReportColumn.id == column_id).first()
        if not column:
            raise ValueError(f"Report column with ID {column_id} not found")
        
        self.db.delete(column)
        self.db.commit()
        
        logger.info(f"Deleted report column {column_id}: {column.column_name}")
    
    def get_report_templates(self, user_id: Optional[int] = None) -> List[models.ReportTemplate]:
        """Get report templates, optionally filtered by user"""
        query = self.db.query(models.ReportTemplate)
        if user_id:
            query = query.filter(
                (models.ReportTemplate.created_by == user_id) |
                (models.ReportTemplate.is_default == True)
            )
        return query.order_by(models.ReportTemplate.is_default.desc(), models.ReportTemplate.name).all()
    
    def create_report_template(self, template_data: Dict[str, Any]) -> models.ReportTemplate:
        """Create a new report template"""
        # Ensure selected_columns is stored as JSON string
        if "selected_columns" in template_data and isinstance(template_data["selected_columns"], list):
            template_data["selected_columns"] = json.dumps(template_data["selected_columns"])
        
        template = models.ReportTemplate(
            **template_data,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc)
        )
        self.db.add(template)
        self.db.commit()
        self.db.refresh(template)
        
        logger.info(f"Created report template: {template.name}")
        return template
    
    def update_report_template(self, template_id: int, update_data: Dict[str, Any]) -> models.ReportTemplate:
        """Update a report template"""
        template = self.db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
        if not template:
            raise ValueError(f"Report template with ID {template_id} not found")
        
        # Ensure selected_columns is stored as JSON string
        if "selected_columns" in update_data and isinstance(update_data["selected_columns"], list):
            update_data["selected_columns"] = json.dumps(update_data["selected_columns"])
        
        for field, value in update_data.items():
            if hasattr(template, field):
                setattr(template, field, value)
        
        template.updated_at = datetime.now(timezone.utc)
        self.db.commit()
        self.db.refresh(template)
        
        logger.info(f"Updated report template {template_id}: {template.name}")
        return template
    
    def delete_report_template(self, template_id: int):
        """Delete a report template"""
        template = self.db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
        if not template:
            raise ValueError(f"Report template with ID {template_id} not found")
        
        self.db.delete(template)
        self.db.commit()
        
        logger.info(f"Deleted report template {template_id}: {template.name}")
    
    def generate_dynamic_excel_report(
        self, 
        selected_columns: List[str], 
        months: int = 2,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> bytes:
        """Generate a dynamic report as Excel file"""
        if not EXCEL_AVAILABLE:
            raise ValueError("Excel functionality not available. Please install openpyxl.")
        
        # Generate report data
        report_data = self.generate_dynamic_report(
            selected_columns=selected_columns,
            months=months,
            start_date=start_date,
            end_date=end_date
        )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dynamic Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Add summary information
        ws.append(["Dynamic Booking Report"])
        ws.append([])
        ws.append(["Period", f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
        ws.append(["Total Bookings", report_data['summary']['total_bookings']])
        ws.append(["Unique Users", report_data['summary']['unique_users']])
        ws.append(["Total Hours", report_data['summary']['total_hours']])
        ws.append(["Average Duration", report_data['summary']['avg_booking_duration']])
        ws.append([])
        
        # Add column headers
        headers = [col["display_label"] for col in report_data["columns"]]
        ws.append(headers)
        
        # Style headers
        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Add data rows
        for record in report_data["data"]:
            row_data = []
            for col in report_data["columns"]:
                col_name = col["column_name"]
                value = record.get(col_name)
                
                # Format value based on data type
                if col["data_type"] == "array" and isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif col["data_type"] == "number" and isinstance(value, (int, float)):
                    value = round(value, 2) if isinstance(value, float) else value
                elif value is None:
                    value = ""
                
                row_data.append(value)
            
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)
        
        # Format summary section
        ws['A1'].font = Font(bold=True, size=14)
        for row in range(3, 8):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.read()
    
    def send_dynamic_report_email(
        self, 
        recipients: List[str], 
        template_id: int,
        months: int = 2,
        include_excel: bool = True
    ) -> bool:
        """Send dynamic report via email using a template"""
        try:
            # Get email settings
            email_settings = self.db.query(models.EmailSettings).first()
            if not email_settings or not email_settings.sendgrid_api_key:
                logger.error("Email settings not configured")
                return False
            
            # Get template
            template = self.db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
            if not template:
                logger.error(f"Report template {template_id} not found")
                return False
            
            # Parse selected columns
            try:
                selected_columns = json.loads(template.selected_columns) if template.selected_columns else []
            except json.JSONDecodeError:
                logger.error(f"Invalid selected_columns format in template {template_id}")
                return False
            
            if not selected_columns:
                logger.error(f"Template {template_id} has no columns configured")
                return False
            
            # Generate report data
            report_data = self.generate_dynamic_report(
                selected_columns=selected_columns,
                months=months
            )
            
            # Import email service here to avoid circular imports
            from .email_service import EmailService
            
            # Create email service
            email_service = EmailService(self.db)
            
            # Generate email content
            subject = f"Dynamic Parking Report - {template.name} - {report_data['period']['start_date']} to {report_data['period']['end_date']}"
            
            html_content = self._generate_dynamic_report_html(report_data, template)
            plain_content = self._generate_dynamic_report_plain_text(report_data, template)
            
            # Prepare attachments
            attachments = []
            if include_excel and EXCEL_AVAILABLE:
                excel_data = self.generate_dynamic_excel_report(
                    selected_columns=selected_columns,
                    months=months
                )
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"dynamic_report_{template.name.replace(' ', '_')}_{timestamp}.xlsx"
                
                attachments.append({
                    'content': excel_data,
                    'filename': filename,
                    'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                })
            
            # Send emails to all recipients
            success_count = 0
            for recipient in recipients:
                try:
                    # Create email data for SendGrid API
                    email_data = {
                        "personalizations": [{
                            "to": [{"email": recipient}],
                            "subject": subject
                        }],
                        "from": {
                            "email": email_settings.from_email,
                            "name": email_settings.from_name
                        },
                        "content": [
                            {
                                "type": "text/plain",
                                "value": plain_content
                            },
                            {
                                "type": "text/html",
                                "value": html_content
                            }
                        ]
                    }
                    
                    # Add attachments if any
                    if attachments:
                        import base64
                        email_data["attachments"] = [
                            {
                                "content": base64.b64encode(attachment['content']).decode('utf-8') if isinstance(attachment['content'], bytes) else attachment['content'],
                                "filename": attachment['filename'],
                                "type": attachment['type'],
                                "disposition": "attachment"
                            }
                            for attachment in attachments
                        ]
                    
                    result = email_service._send_email_request(email_data)
                    
                    if result['success']:
                        success_count += 1
                        logger.info(f"Dynamic report sent successfully to {recipient}")
                    else:
                        logger.error(f"Failed to send dynamic report to {recipient}: {result.get('error', 'Unknown error')}")
                        
                except Exception as e:
                    logger.error(f"Error sending dynamic report to {recipient}: {str(e)}")
            
            if success_count > 0:
                logger.info(f"Dynamic report sent to {success_count}/{len(recipients)} recipients")
                return True
            else:
                logger.error("Failed to send dynamic report to any recipients")
                return False
                
        except Exception as e:
            logger.error(f"Error sending dynamic report: {str(e)}")
            return False
    
    def send_scheduled_dynamic_report(self, force_send: bool = False) -> bool:
        """Send scheduled dynamic report to configured recipients"""
        try:
            email_settings = self.db.query(models.EmailSettings).first()
            if not email_settings or not email_settings.dynamic_reports_enabled:
                logger.info("Dynamic reports are disabled")
                return False
            
            if not email_settings.dynamic_report_recipients:
                logger.info("No dynamic report recipients configured")
                return False
            
            if not email_settings.dynamic_report_template_id:
                logger.info("No dynamic report template configured")
                return False
            
            if not email_settings.sendgrid_api_key:
                logger.error("SendGrid API key not configured")
                return False
            
            # Parse recipients from JSON string
            try:
                recipients = json.loads(email_settings.dynamic_report_recipients) if isinstance(email_settings.dynamic_report_recipients, str) else email_settings.dynamic_report_recipients
            except (json.JSONDecodeError, TypeError):
                logger.error("Invalid dynamic report recipients format")
                return False
            
            if not recipients:
                logger.info("No dynamic report recipients configured")
                return False
            
            # Check if we should send report based on schedule
            now = datetime.now(timezone.utc)
            if not force_send and email_settings.last_dynamic_report_sent:
                time_since_last = now - email_settings.last_dynamic_report_sent
                
                if email_settings.dynamic_report_frequency == "daily" and time_since_last < timedelta(hours=23):
                    logger.info("Daily dynamic report already sent recently")
                    return False
                elif email_settings.dynamic_report_frequency == "weekly" and time_since_last < timedelta(days=6):
                    logger.info("Weekly dynamic report already sent recently")
                    return False
                elif email_settings.dynamic_report_frequency == "monthly" and time_since_last < timedelta(days=29):
                    logger.info("Monthly dynamic report already sent recently")
                    return False
            
            # Send the report
            success = self.send_dynamic_report_email(
                recipients=recipients,
                template_id=email_settings.dynamic_report_template_id,
                months=2,  # Default to 2 months
                include_excel=True
            )
            
            if success:
                # Update last report sent time
                email_settings.last_dynamic_report_sent = now
                self.db.commit()
                logger.info("Scheduled dynamic report sent successfully")
                return True
            else:
                logger.error("Failed to send scheduled dynamic report")
                return False
                
        except Exception as e:
            logger.error(f"Error sending scheduled dynamic report: {str(e)}")
            return False
    
    def _generate_dynamic_report_html(self, report_data: Dict[str, Any], template: models.ReportTemplate) -> str:
        """Generate HTML content for dynamic report email"""
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Dynamic Parking Report - {template.name}</h2>
            <p><strong>Report Period:</strong> {report_data['period']['start_date']} to {report_data['period']['end_date']}</p>
            
            <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3>Summary</h3>
                <ul>
                    <li><strong>Total Bookings:</strong> {report_data['summary']['total_bookings']}</li>
                    <li><strong>Unique Users:</strong> {report_data['summary']['unique_users']}</li>
                    <li><strong>Total Hours:</strong> {report_data['summary']['total_hours']}</li>
                    <li><strong>Average Booking Duration:</strong> {report_data['summary']['avg_booking_duration']} hours</li>
                </ul>
            </div>
            
            <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h3>Report Data (Top 20 Records)</h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr style="background-color: #e0e0e0;">
        """
        
        # Add column headers
        for col in report_data['columns']:
            html_content += f'<th style="border: 1px solid #ccc; padding: 8px; text-align: left;">{col["display_label"]}</th>'
        
        html_content += "</tr>"
        
        # Add data rows (limit to top 20)
        for i, record in enumerate(report_data['data'][:20]):
            html_content += "<tr>"
            for col in report_data['columns']:
                value = record.get(col['column_name'], '')
                if col['data_type'] == 'array' and isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif col['data_type'] == 'number' and isinstance(value, (int, float)):
                    value = round(value, 2) if isinstance(value, float) else value
                elif value is None:
                    value = ''
                
                html_content += f'<td style="border: 1px solid #ccc; padding: 8px;">{value}</td>'
            html_content += "</tr>"
        
        if len(report_data['data']) > 20:
            html_content += f"""
                    <tr>
                        <td colspan="{len(report_data['columns'])}" style="border: 1px solid #ccc; padding: 8px; text-align: center; font-style: italic;">
                            ... and {len(report_data['data']) - 20} more records (see attached Excel file for complete data)
                        </td>
                    </tr>
            """
        
        html_content += """
                </table>
            </div>
            
            <p><em>For detailed statistics and complete data, please see the attached Excel report.</em></p>
            
            <p>This report was generated automatically by the parking booking system.</p>
            
            <p>Best regards,<br>Parking Booking System</p>
        </body>
        </html>
        """
        
        return html_content
    
    def _generate_dynamic_report_plain_text(self, report_data: Dict[str, Any], template: models.ReportTemplate) -> str:
        """Generate plain text content for dynamic report email"""
        plain_content = f"""
        Dynamic Parking Report - {template.name}
        Report Period: {report_data['period']['start_date']} to {report_data['period']['end_date']}
        
        SUMMARY
        Total Bookings: {report_data['summary']['total_bookings']}
        Unique Users: {report_data['summary']['unique_users']}
        Total Hours: {report_data['summary']['total_hours']}
        Average Booking Duration: {report_data['summary']['avg_booking_duration']} hours
        
        REPORT DATA (Top 20 Records)
        """
        
        # Add column headers
        headers = [col['display_label'] for col in report_data['columns']]
        plain_content += ' | '.join(headers) + '\n'
        plain_content += '-' * (sum(len(h) for h in headers) + len(headers) * 3 - 3) + '\n'
        
        # Add data rows (limit to top 20)
        for i, record in enumerate(report_data['data'][:20]):
            row_values = []
            for col in report_data['columns']:
                value = record.get(col['column_name'], '')
                if col['data_type'] == 'array' and isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                elif col['data_type'] == 'number' and isinstance(value, (int, float)):
                    value = round(value, 2) if isinstance(value, float) else value
                elif value is None:
                    value = ''
                
                row_values.append(str(value))
            
            plain_content += ' | '.join(row_values) + '\n'
        
        if len(report_data['data']) > 20:
            plain_content += f"... and {len(report_data['data']) - 20} more records (see attached Excel file for complete data)\n"
        
        plain_content += """
        
        For detailed statistics and complete data, please see the attached Excel report.
        
        This report was generated automatically by the parking booking system.
        
        Best regards,
        Parking Booking System
        """
        
        return plain_content
