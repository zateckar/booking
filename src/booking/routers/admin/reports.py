from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, extract
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any
import io
import json

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ... import models, schemas
from ...database import get_db
from ...email_service import EmailService
from ...timezone_service import TimezoneService
from ...security import get_current_admin_user

router = APIRouter()


@router.get("/reports/bookings")
def get_booking_reports(
    months: int = 2,  # Current and previous month by default
    current_admin_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Get booking reports for specified number of months"""
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    start_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    if months == 0:
        # Special case: Last month only
        # Start of last month
        if start_of_current_month.month == 1:
            start_date = start_of_current_month.replace(year=start_of_current_month.year - 1, month=12)
        else:
            start_date = start_of_current_month.replace(month=start_of_current_month.month - 1)
        
        # End of last month (start of current month)
        end_date = start_of_current_month
    else:
        # Original logic: current month + previous months
        start_date = start_of_current_month - timedelta(days=32 * (months - 1))  # Rough calculation
        start_date = start_date.replace(day=1)  # First day of the target month
        end_date = now
    
    # Query bookings in the date range
    bookings = db.query(models.Booking).join(models.User).join(models.ParkingSpace).join(models.ParkingLot).filter(
        and_(
            models.Booking.start_time >= start_date,
            models.Booking.start_time <= end_date,
            models.Booking.is_cancelled == False
        )
    ).all()
    
    # Calculate statistics
    total_bookings = len(bookings)
    unique_users = len(set(booking.user_id for booking in bookings))
    
    # User statistics
    user_stats = {}
    parking_lot_stats = {}
    monthly_stats = {}
    
    for booking in bookings:
        # User statistics
        user_email = booking.user.email
        if user_email not in user_stats:
            user_stats[user_email] = {
                'user_id': booking.user_id,
                'email': user_email,
                'total_bookings': 0,
                'total_hours': 0,
                'avg_duration': 0,
                'parking_lots_used': set(),
                'license_plates': set()
            }
        
        duration_hours = (booking.end_time - booking.start_time).total_seconds() / 3600
        user_stats[user_email]['total_bookings'] += 1
        user_stats[user_email]['total_hours'] += duration_hours
        user_stats[user_email]['parking_lots_used'].add(booking.space.parking_lot.name)
        user_stats[user_email]['license_plates'].add(booking.license_plate)
        
        # Parking lot statistics
        lot_name = booking.space.parking_lot.name
        if lot_name not in parking_lot_stats:
            parking_lot_stats[lot_name] = {
                'name': lot_name,
                'total_bookings': 0,
                'total_hours': 0,
                'unique_users': set()
            }
        
        parking_lot_stats[lot_name]['total_bookings'] += 1
        parking_lot_stats[lot_name]['total_hours'] += duration_hours
        parking_lot_stats[lot_name]['unique_users'].add(booking.user_id)
        
        # Monthly statistics
        month_key = booking.start_time.strftime('%Y-%m')
        if month_key not in monthly_stats:
            monthly_stats[month_key] = {
                'month': month_key,
                'total_bookings': 0,
                'total_hours': 0,
                'unique_users': set(),
                'avg_duration': 0
            }
        
        monthly_stats[month_key]['total_bookings'] += 1
        monthly_stats[month_key]['total_hours'] += duration_hours
        monthly_stats[month_key]['unique_users'].add(booking.user_id)
    
    # Calculate averages and convert sets to counts
    for user_email, stats in user_stats.items():
        if stats['total_bookings'] > 0:
            stats['avg_duration'] = stats['total_hours'] / stats['total_bookings']
        stats['parking_lots_used'] = len(stats['parking_lots_used'])
        stats['license_plates'] = len(stats['license_plates'])
    
    for lot_name, stats in parking_lot_stats.items():
        stats['unique_users'] = len(stats['unique_users'])
        if stats['total_bookings'] > 0:
            stats['avg_duration'] = stats['total_hours'] / stats['total_bookings']
    
    for month_key, stats in monthly_stats.items():
        stats['unique_users'] = len(stats['unique_users'])
        if stats['total_bookings'] > 0:
            stats['avg_duration'] = stats['total_hours'] / stats['total_bookings']
    
    return {
        'period': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'months': months
        },
        'summary': {
            'total_bookings': total_bookings,
            'unique_users': unique_users,
            'total_hours': sum(stats['total_hours'] for stats in user_stats.values()),
            'avg_booking_duration': sum(stats['total_hours'] for stats in user_stats.values()) / total_bookings if total_bookings > 0 else 0
        },
        'user_statistics': list(user_stats.values()),
        'parking_lot_statistics': list(parking_lot_stats.values()),
        'monthly_statistics': list(monthly_stats.values()),
        'raw_bookings': [
            {
                'id': booking.id,
                'user_email': booking.user.email,
                'start_time': booking.start_time.isoformat(),
                'end_time': booking.end_time.isoformat(),
                'duration_hours': (booking.end_time - booking.start_time).total_seconds() / 3600,
                'license_plate': booking.license_plate,
                'parking_lot': booking.space.parking_lot.name,
                'space_number': booking.space.space_number
            }
            for booking in bookings
        ]
    }


@router.get("/reports/download/excel")
def download_excel_report(
    months: int = 2,
    current_admin_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Download booking report as Excel file"""
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel functionality not available. Please install openpyxl.")
    
    # Get report data
    report_data = get_booking_reports(months=months, db=db)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Summary", 0)
    user_stats_sheet = wb.create_sheet("User Statistics", 1)
    lot_stats_sheet = wb.create_sheet("Parking Lot Statistics", 2)
    monthly_stats_sheet = wb.create_sheet("Monthly Statistics", 3)
    raw_data_sheet = wb.create_sheet("Raw Data", 4)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Summary Sheet
    summary_sheet.append(["Booking Report Summary"])
    summary_sheet.append([])
    summary_sheet.append(["Period", f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
    summary_sheet.append(["Total Bookings", report_data['summary']['total_bookings']])
    summary_sheet.append(["Unique Users", report_data['summary']['unique_users']])
    summary_sheet.append(["Total Hours", round(report_data['summary']['total_hours'], 2)])
    summary_sheet.append(["Average Booking Duration (hours)", round(report_data['summary']['avg_booking_duration'], 2)])
    
    # Format summary sheet
    summary_sheet['A1'].font = Font(bold=True, size=14)
    for row in range(3, 8):
        summary_sheet[f'A{row}'].font = Font(bold=True)
    
    # User Statistics Sheet
    if report_data['user_statistics']:
        headers = ["Email", "Total Bookings", "Total Hours", "Average Duration (hours)", "Parking Lots Used", "License Plates Used"]
        user_stats_sheet.append(headers)
        
        for stats in report_data['user_statistics']:
            user_stats_sheet.append([
                stats['email'],
                stats['total_bookings'],
                round(stats['total_hours'], 2),
                round(stats['avg_duration'], 2),
                stats['parking_lots_used'],
                stats['license_plates']
            ])
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = user_stats_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # Parking Lot Statistics Sheet
    if report_data['parking_lot_statistics']:
        headers = ["Parking Lot", "Total Bookings", "Total Hours", "Average Duration (hours)", "Unique Users"]
        lot_stats_sheet.append(headers)
        
        for stats in report_data['parking_lot_statistics']:
            lot_stats_sheet.append([
                stats['name'],
                stats['total_bookings'],
                round(stats['total_hours'], 2),
                round(stats.get('avg_duration', 0), 2),
                stats['unique_users']
            ])
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = lot_stats_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # Monthly Statistics Sheet
    if report_data['monthly_statistics']:
        headers = ["Month", "Total Bookings", "Total Hours", "Average Duration (hours)", "Unique Users"]
        monthly_stats_sheet.append(headers)
        
        for stats in report_data['monthly_statistics']:
            monthly_stats_sheet.append([
                stats['month'],
                stats['total_bookings'],
                round(stats['total_hours'], 2),
                round(stats['avg_duration'], 2),
                stats['unique_users']
            ])
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = monthly_stats_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # Raw Data Sheet
    if report_data['raw_bookings']:
        headers = ["Booking ID", "User Email", "Start Time", "End Time", "Duration (hours)", "License Plate", "Parking Lot", "Space Number"]
        raw_data_sheet.append(headers)
        
        for booking in report_data['raw_bookings']:
            raw_data_sheet.append([
                booking['id'],
                booking['user_email'],
                booking['start_time'],
                booking['end_time'],
                round(booking['duration_hours'], 2),
                booking['license_plate'],
                booking['parking_lot'],
                booking['space_number']
            ])
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = raw_data_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename with current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"booking_report_{timestamp}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/reports/send-email")
def send_report_email(
    recipients: List[str],
    months: int = 2,
    include_excel: bool = True,
    current_admin_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Send booking report via email"""
    
    try:
        # Get email settings
        email_settings = db.query(models.EmailSettings).first()
        if not email_settings or not email_settings.sendgrid_api_key:
            raise HTTPException(status_code=400, detail="Email settings not configured")
        
        # Get report data
        report_data = get_booking_reports(months=months, db=db)
        
        # Create email service
        email_service = EmailService(
            api_key=email_settings.sendgrid_api_key,
            from_email=email_settings.from_email,
            from_name=email_settings.from_name
        )
        
        # Generate email content
        subject = f"Parking Booking Report - {report_data['period']['start_date']} to {report_data['period']['end_date']}"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Parking Booking Report</h2>
            
            <h3>Summary</h3>
            <ul>
                <li><strong>Period:</strong> {report_data['period']['start_date']} to {report_data['period']['end_date']}</li>
                <li><strong>Total Bookings:</strong> {report_data['summary']['total_bookings']}</li>
                <li><strong>Unique Users:</strong> {report_data['summary']['unique_users']}</li>
                <li><strong>Total Hours:</strong> {round(report_data['summary']['total_hours'], 2)}</li>
                <li><strong>Average Booking Duration:</strong> {round(report_data['summary']['avg_booking_duration'], 2)} hours</li>
            </ul>
            
            <h3>Top Users</h3>
            <table style="border-collapse: collapse; width: 100%;">
                <tr style="background-color: #f2f2f2;">
                    <th style="border: 1px solid #ddd; padding: 8px;">Email</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">Total Bookings</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">Total Hours</th>
                </tr>
        """
        
        # Add top 10 users
        sorted_users = sorted(report_data['user_statistics'], key=lambda x: x['total_bookings'], reverse=True)[:10]
        for user in sorted_users:
            html_content += f"""
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">{user['email']}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{user['total_bookings']}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{round(user['total_hours'], 2)}</td>
                </tr>
            """
        
        html_content += """
            </table>
            
            <h3>Parking Lot Usage</h3>
            <table style="border-collapse: collapse; width: 100%;">
                <tr style="background-color: #f2f2f2;">
                    <th style="border: 1px solid #ddd; padding: 8px;">Parking Lot</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">Total Bookings</th>
                    <th style="border: 1px solid #ddd; padding: 8px;">Unique Users</th>
                </tr>
        """
        
        for lot in report_data['parking_lot_statistics']:
            html_content += f"""
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">{lot['name']}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{lot['total_bookings']}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{lot['unique_users']}</td>
                </tr>
            """
        
        html_content += """
            </table>
            
            <p><em>For detailed statistics, please see the attached Excel report.</em></p>
            
            <p>Best regards,<br>Parking Booking System</p>
        </body>
        </html>
        """
        
        # Prepare attachments
        attachments = []
        if include_excel and EXCEL_AVAILABLE:
            # Generate Excel file
            report_data_for_excel = get_booking_reports(months=months, db=db)
            
            # Create Excel workbook (simplified version for email)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Booking Report"
            
            # Add summary
            ws.append(["Booking Report Summary"])
            ws.append([])
            ws.append(["Total Bookings", report_data['summary']['total_bookings']])
            ws.append(["Unique Users", report_data['summary']['unique_users']])
            ws.append(["Total Hours", round(report_data['summary']['total_hours'], 2)])
            ws.append([])
            
            # Add user statistics
            ws.append(["User Statistics"])
            ws.append(["Email", "Total Bookings", "Total Hours", "Average Duration"])
            for user in report_data['user_statistics']:
                ws.append([user['email'], user['total_bookings'], round(user['total_hours'], 2), round(user['avg_duration'], 2)])
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"booking_report_{timestamp}.xlsx"
            
            attachments.append({
                'content': excel_buffer.read(),
                'filename': filename,
                'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
        
        # Send email
        success = email_service.send_email(
            to_emails=recipients,
            subject=subject,
            html_content=html_content,
            attachments=attachments
        )
        
        if success:
            return {"message": "Report sent successfully", "recipients": recipients}
        else:
            raise HTTPException(status_code=500, detail="Failed to send email")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error sending report: {str(e)}")


@router.get("/reports/schedule-settings")
def get_schedule_settings(
    current_admin_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Get current report scheduling settings"""
    email_settings = db.query(models.EmailSettings).first()
    if not email_settings:
        return {
            "reports_enabled": False,
            "report_recipients": [],
            "report_schedule_hour": 9,
            "report_frequency": "daily",
            "last_report_sent": None,
            "timezone": "UTC"
        }
    
    # Parse report recipients from JSON string
    try:
        recipients = json.loads(email_settings.report_recipients) if email_settings.report_recipients else []
    except:
        recipients = []
    
    return {
        "reports_enabled": email_settings.reports_enabled,
        "report_recipients": recipients,
        "report_schedule_hour": email_settings.report_schedule_hour,
        "report_frequency": email_settings.report_frequency,
        "last_report_sent": email_settings.last_report_sent.isoformat() if email_settings.last_report_sent else None,
        "timezone": email_settings.timezone
    }


@router.put("/reports/schedule-settings")
def update_schedule_settings(
    settings: schemas.EmailSettingsUpdate,
    current_admin_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db)
):
    """Update report scheduling settings"""
    
    email_settings = db.query(models.EmailSettings).first()
    if not email_settings:
        # Create new settings record
        email_settings = models.EmailSettings()
        db.add(email_settings)
    
    # Update fields if provided
    if settings.reports_enabled is not None:
        email_settings.reports_enabled = settings.reports_enabled
    if settings.report_recipients is not None:
        email_settings.report_recipients = json.dumps(settings.report_recipients)
    if settings.report_schedule_hour is not None:
        email_settings.report_schedule_hour = settings.report_schedule_hour
    if settings.report_frequency is not None:
        email_settings.report_frequency = settings.report_frequency
    if settings.timezone is not None:
        email_settings.timezone = settings.timezone
    
    db.commit()
    db.refresh(email_settings)
    
    return {"message": "Schedule settings updated successfully"}
