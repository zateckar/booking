"""
Admin endpoints for dynamic reporting with configurable columns
"""

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from datetime import datetime, timezone
import json
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ... import models, schemas
from ...database import get_db
from ...security import get_current_admin_user
from ...dynamic_reports_service import DynamicReportsService
from ...logging_config import get_logger

router = APIRouter()
logger = get_logger("dynamic_reports_admin")


@router.get("/columns")
def get_available_columns(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Get all available columns for dynamic reports"""
    try:
        reports_service = DynamicReportsService(db)
        columns = reports_service.get_available_columns()
        
        logger.info(f"Retrieved {len(columns)} available report columns")
        return {"columns": columns}
        
    except Exception as e:
        logger.error(f"Error retrieving available columns: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving available columns: {str(e)}")


@router.post("/generate")
def generate_dynamic_report(
    request: schemas.DynamicReportRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Generate a dynamic report with selected columns"""
    try:
        reports_service = DynamicReportsService(db)
        
        report_data = reports_service.generate_dynamic_report(
            selected_columns=request.selected_columns,
            months=request.months,
            start_date=request.start_date,
            end_date=request.end_date
        )
        
        date_info = ""
        if request.start_date and request.end_date:
            date_info = f" from {request.start_date.date()} to {request.end_date.date()}"
        else:
            date_info = f" for {request.months} months"
        
        logger.info(f"Generated dynamic report with {len(request.selected_columns)} columns{date_info}")
        return report_data
        
    except Exception as e:
        logger.error(f"Error generating dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating dynamic report: {str(e)}")


@router.post("/generate/excel")
def generate_dynamic_excel_report(
    request: schemas.DynamicReportRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Generate a dynamic report as Excel file"""
    try:
        if not EXCEL_AVAILABLE:
            raise HTTPException(status_code=500, detail="Excel functionality not available. Please install openpyxl.")
        
        reports_service = DynamicReportsService(db)
        
        # Generate report data
        report_data = reports_service.generate_dynamic_report(
            selected_columns=request.selected_columns,
            months=request.months,
            start_date=request.start_date,
            end_date=request.end_date
        )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dynamic Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Add summary information
        ws.append(["Dynamic Booking Report"])
        ws.append([])
        ws.append(["Period", f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
        ws.append(["Total Bookings", report_data['summary']['total_bookings']])
        ws.append(["Unique Users", report_data['summary']['unique_users']])
        ws.append(["Total Hours", report_data['summary']['total_hours']])
        ws.append(["Average Duration", report_data['summary']['avg_booking_duration']])
        ws.append([])
        
        # Add column headers
        headers = [col["display_label"] for col in report_data["columns"]]
        ws.append(headers)
        
        # Style headers
        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Add data rows
        for record in report_data["data"]:
            row_data = []
            for col in report_data["columns"]:
                col_name = col["column_name"]
                value = record.get(col_name)
                
                # Format value based on data type
                if col["data_type"] == "array" and isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif col["data_type"] == "number" and isinstance(value, (int, float)):
                    value = round(value, 2) if isinstance(value, float) else value
                elif value is None:
                    value = ""
                
                row_data.append(value)
            
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)
        
        # Format summary section
        ws['A1'].font = Font(bold=True, size=14)
        for row in range(3, 8):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename with current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dynamic_report_{timestamp}.xlsx"
        
        logger.info(f"Generated dynamic Excel report with {len(request.selected_columns)} columns")
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error generating dynamic Excel report: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating dynamic Excel report: {str(e)}")


@router.post("/columns", response_model=schemas.ReportColumn)
def create_report_column(
    column_data: schemas.ReportColumnCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Create a new report column configuration"""
    try:
        reports_service = DynamicReportsService(db)
        column = reports_service.create_report_column(column_data.model_dump())
        
        logger.info(f"Created report column: {column.column_name}")
        return column
        
    except Exception as e:
        logger.error(f"Error creating report column: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating report column: {str(e)}")


@router.put("/columns/{column_id}", response_model=schemas.ReportColumn)
def update_report_column(
    column_id: int,
    column_data: schemas.ReportColumnUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Update a report column configuration"""
    try:
        reports_service = DynamicReportsService(db)
        
        update_dict = {k: v for k, v in column_data.model_dump().items() if v is not None}
        column = reports_service.update_report_column(column_id, update_dict)
        
        logger.info(f"Updated report column {column_id}: {column.column_name}")
        return column
        
    except ValueError as e:
        logger.warning(f"Report column not found: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error updating report column: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating report column: {str(e)}")


@router.delete("/columns/{column_id}")
def delete_report_column(
    column_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Delete a report column configuration"""
    try:
        reports_service = DynamicReportsService(db)
        reports_service.delete_report_column(column_id)
        
        logger.info(f"Deleted report column {column_id}")
        return {"message": f"Report column {column_id} deleted successfully"}
        
    except ValueError as e:
        logger.warning(f"Report column not found: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error deleting report column: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting report column: {str(e)}")


@router.get("/templates", response_model=List[schemas.ReportTemplate])
def get_report_templates(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Get all report templates available to the current user"""
    try:
        reports_service = DynamicReportsService(db)
        templates = reports_service.get_report_templates(user_id=current_user.id)
        
        # Convert selected_columns from JSON string to list for response
        result = []
        for template in templates:
            template_dict = {
                "id": template.id,
                "name": template.name,
                "description": template.description,
                "selected_columns": [],
                "is_default": template.is_default,
                "created_by": template.created_by,
                "created_at": template.created_at,
                "updated_at": template.updated_at
            }
            
            if template.selected_columns:
                try:
                    template_dict["selected_columns"] = json.loads(template.selected_columns)
                except json.JSONDecodeError:
                    template_dict["selected_columns"] = []
            
            result.append(schemas.ReportTemplate(**template_dict))
        
        logger.info(f"Retrieved {len(result)} report templates for user {current_user.id}")
        return result
        
    except Exception as e:
        logger.error(f"Error retrieving report templates: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving report templates: {str(e)}")


@router.post("/templates", response_model=schemas.ReportTemplate)
def create_report_template(
    template_data: schemas.ReportTemplateCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Create a new report template"""
    try:
        reports_service = DynamicReportsService(db)
        
        template_dict = template_data.model_dump()
        template_dict["created_by"] = current_user.id
        
        template = reports_service.create_report_template(template_dict)
        
        # Convert for response
        response_dict = {
            "id": template.id,
            "name": template.name,
            "description": template.description,
            "selected_columns": template_data.selected_columns,  # Use original list
            "is_default": template.is_default,
            "created_by": template.created_by,
            "created_at": template.created_at,
            "updated_at": template.updated_at
        }
        
        logger.info(f"Created report template: {template.name}")
        return schemas.ReportTemplate(**response_dict)
        
    except Exception as e:
        logger.error(f"Error creating report template: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating report template: {str(e)}")


@router.put("/templates/{template_id}", response_model=schemas.ReportTemplate)
def update_report_template(
    template_id: int,
    template_data: schemas.ReportTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Update a report template"""
    try:
        reports_service = DynamicReportsService(db)
        
        # Check if user owns template or is updating a default template
        template = db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Report template not found")
        
        if template.created_by != current_user.id and not template.is_default:
            raise HTTPException(status_code=403, detail="You can only modify your own templates")
        
        update_dict = {k: v for k, v in template_data.model_dump().items() if v is not None}
        updated_template = reports_service.update_report_template(template_id, update_dict)
        
        # Convert for response
        selected_columns = []
        if updated_template.selected_columns:
            try:
                selected_columns = json.loads(updated_template.selected_columns)
            except json.JSONDecodeError:
                selected_columns = []
        
        response_dict = {
            "id": updated_template.id,
            "name": updated_template.name,
            "description": updated_template.description,
            "selected_columns": selected_columns,
            "is_default": updated_template.is_default,
            "created_by": updated_template.created_by,
            "created_at": updated_template.created_at,
            "updated_at": updated_template.updated_at
        }
        
        logger.info(f"Updated report template {template_id}: {updated_template.name}")
        return schemas.ReportTemplate(**response_dict)
        
    except ValueError as e:
        logger.warning(f"Report template not found: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error updating report template: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating report template: {str(e)}")


@router.delete("/templates/{template_id}")
def delete_report_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Delete a report template"""
    try:
        # Check if user owns template
        template = db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Report template not found")
        
        if template.created_by != current_user.id:
            raise HTTPException(status_code=403, detail="You can only delete your own templates")
        
        reports_service = DynamicReportsService(db)
        reports_service.delete_report_template(template_id)
        
        logger.info(f"Deleted report template {template_id}")
        return {"message": f"Report template {template_id} deleted successfully"}
        
    except ValueError as e:
        logger.warning(f"Report template not found: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Error deleting report template: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting report template: {str(e)}")


@router.post("/templates/{template_id}/generate")
def generate_report_from_template(
    template_id: int,
    months: int = 2,
    include_excel: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Generate a report using a saved template"""
    try:
        # Get template
        template = db.query(models.ReportTemplate).filter(models.ReportTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Report template not found")
        
        # Check access
        if template.created_by != current_user.id and not template.is_default:
            raise HTTPException(status_code=403, detail="You can only use your own templates or default templates")
        
        # Parse selected columns
        try:
            selected_columns = json.loads(template.selected_columns) if template.selected_columns else []
        except json.JSONDecodeError:
            selected_columns = []
        
        if not selected_columns:
            raise HTTPException(status_code=400, detail="Template has no columns configured")
        
        # Generate report
        reports_service = DynamicReportsService(db)
        
        if include_excel:
            # Create DynamicReportRequest for Excel generation
            request = schemas.DynamicReportRequest(
                selected_columns=selected_columns,
                months=months,
                include_excel=True
            )
            return generate_dynamic_excel_report(request, db, current_user)
        else:
            report_data = reports_service.generate_dynamic_report(
                selected_columns=selected_columns,
                months=months
            )
            
            # Add template info
            report_data["template"] = {
                "id": template.id,
                "name": template.name,
                "description": template.description
            }
            
        logger.info(f"Generated report from template {template_id}: {template.name}")
        return report_data
        
    except Exception as e:
        logger.error(f"Error generating report from template: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating report from template: {str(e)}")


@router.post("/send-email")
def send_dynamic_report_email(
    request: Dict[str, Any],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Send dynamic report via email using a template"""
    try:
        # Extract parameters from request body
        template_id = request.get("template_id")
        recipients = request.get("recipients", [])
        months = request.get("months", 2)
        include_excel = request.get("include_excel", True)
        
        # Validate required parameters
        if not template_id:
            raise HTTPException(status_code=400, detail="Template ID is required")
        if not recipients:
            raise HTTPException(status_code=400, detail="Recipients are required")
        
        reports_service = DynamicReportsService(db)
        
        success = reports_service.send_dynamic_report_email(
            recipients=recipients,
            template_id=template_id,
            months=months,
            include_excel=include_excel
        )
        
        if success:
            logger.info(f"Dynamic report sent successfully to {len(recipients)} recipients")
            return {"message": "Dynamic report sent successfully", "recipients": recipients}
        else:
            raise HTTPException(status_code=500, detail="Failed to send dynamic report")
            
    except Exception as e:
        logger.error(f"Error sending dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error sending dynamic report: {str(e)}")


@router.get("/schedule-settings")
def get_dynamic_report_schedule_settings(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Get current dynamic report scheduling settings"""
    try:
        email_settings = db.query(models.EmailSettings).first()
        if not email_settings:
            return {
                "dynamic_reports_enabled": False,
                "dynamic_report_recipients": [],
                "dynamic_report_schedule_hour": 9,
                "dynamic_report_frequency": "weekly",
                "dynamic_report_template_id": None,
                "last_dynamic_report_sent": None,
                "timezone": "UTC"
            }
        
        # Parse report recipients from JSON string
        try:
            recipients = json.loads(email_settings.dynamic_report_recipients) if email_settings.dynamic_report_recipients else []
        except:
            recipients = []
        
        result = {
            "dynamic_reports_enabled": email_settings.dynamic_reports_enabled or False,
            "dynamic_report_recipients": recipients,
            "dynamic_report_schedule_hour": email_settings.dynamic_report_schedule_hour or 9,
            "dynamic_report_frequency": email_settings.dynamic_report_frequency or "weekly",
            "dynamic_report_template_id": email_settings.dynamic_report_template_id,
            "last_dynamic_report_sent": email_settings.last_dynamic_report_sent.isoformat() if email_settings.last_dynamic_report_sent else None,
            "timezone": email_settings.timezone or "UTC"
        }
        
        logger.info("Retrieved dynamic report schedule settings")
        return result
        
    except Exception as e:
        logger.error(f"Error retrieving dynamic report schedule settings: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving schedule settings: {str(e)}")


@router.put("/schedule-settings")
def update_dynamic_report_schedule_settings(
    settings: schemas.EmailSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Update dynamic report scheduling settings"""
    try:
        email_settings = db.query(models.EmailSettings).first()
        if not email_settings:
            # Create new settings record
            email_settings = models.EmailSettings()
            db.add(email_settings)
        
        # Update dynamic report fields if provided
        if settings.dynamic_reports_enabled is not None:
            email_settings.dynamic_reports_enabled = settings.dynamic_reports_enabled
        if settings.dynamic_report_recipients is not None:
            email_settings.dynamic_report_recipients = json.dumps(settings.dynamic_report_recipients)
        if settings.dynamic_report_schedule_hour is not None:
            email_settings.dynamic_report_schedule_hour = settings.dynamic_report_schedule_hour
        if settings.dynamic_report_frequency is not None:
            email_settings.dynamic_report_frequency = settings.dynamic_report_frequency
        if settings.dynamic_report_template_id is not None:
            email_settings.dynamic_report_template_id = settings.dynamic_report_template_id
        if settings.timezone is not None:
            email_settings.timezone = settings.timezone
        
        db.commit()
        db.refresh(email_settings)
        
        logger.info("Updated dynamic report schedule settings")
        return {"message": "Dynamic report schedule settings updated successfully"}
        
    except Exception as e:
        logger.error(f"Error updating dynamic report schedule settings: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating schedule settings: {str(e)}")


@router.post("/send-test-email")
def send_test_dynamic_report(
    request: Dict[str, Any],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Send a test dynamic report to a single email address"""
    try:
        # Extract parameters from request body
        template_id = request.get("template_id")
        test_email = request.get("test_email")
        
        # Validate required parameters
        if not template_id:
            raise HTTPException(status_code=400, detail="Template ID is required")
        if not test_email:
            raise HTTPException(status_code=400, detail="Test email address is required")
        
        reports_service = DynamicReportsService(db)
        
        success = reports_service.send_dynamic_report_email(
            recipients=[test_email],
            template_id=template_id,
            months=1,  # Use 1 month for test to reduce data
            include_excel=True
        )
        
        if success:
            logger.info(f"Test dynamic report sent successfully to {test_email}")
            return {"message": f"Test dynamic report sent successfully to {test_email}"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send test dynamic report")
            
    except Exception as e:
        logger.error(f"Error sending test dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error sending test dynamic report: {str(e)}")


# Scheduled Dynamic Reports Endpoints
@router.get("/schedules", response_model=List[schemas.ScheduledDynamicReport])
def get_scheduled_dynamic_reports(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Get all scheduled dynamic reports"""
    try:
        schedules = db.query(models.ScheduledDynamicReport).filter(
            models.ScheduledDynamicReport.created_by == current_user.id
        ).all()
        
        # Convert to response format, parsing JSON fields
        result = []
        for schedule in schedules:
            # Parse recipients from JSON string
            try:
                recipients = json.loads(schedule.recipients) if schedule.recipients else []
            except json.JSONDecodeError:
                recipients = []
            
            # Get template and parse its selected_columns
            template_data = None
            if schedule.template:
                try:
                    selected_columns = json.loads(schedule.template.selected_columns) if schedule.template.selected_columns else []
                except json.JSONDecodeError:
                    selected_columns = []
                
                template_data = schemas.ReportTemplate(
                    id=schedule.template.id,
                    name=schedule.template.name,
                    description=schedule.template.description,
                    selected_columns=selected_columns,
                    is_default=schedule.template.is_default,
                    created_by=schedule.template.created_by,
                    created_at=schedule.template.created_at,
                    updated_at=schedule.template.updated_at
                )
            
            schedule_data = schemas.ScheduledDynamicReport(
                id=schedule.id,
                name=schedule.name,
                description=schedule.description,
                template_id=schedule.template_id,
                recipients=recipients,
                frequency=schedule.frequency,
                schedule_hour=schedule.schedule_hour,
                timezone=schedule.timezone,
                is_enabled=schedule.is_enabled,
                include_excel=schedule.include_excel,
                months_period=schedule.months_period,
                last_sent=schedule.last_sent,
                last_error=schedule.last_error,
                created_by=schedule.created_by,
                created_at=schedule.created_at,
                updated_at=schedule.updated_at,
                template=template_data
            )
            result.append(schedule_data)
        
        logger.info(f"Retrieved {len(result)} scheduled dynamic reports for user {current_user.id}")
        return result
        
    except Exception as e:
        logger.error(f"Error retrieving scheduled dynamic reports: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving scheduled reports: {str(e)}")


@router.post("/schedules", response_model=schemas.ScheduledDynamicReport)
def create_scheduled_dynamic_report(
    schedule_data: schemas.ScheduledDynamicReportCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Create a new scheduled dynamic report"""
    try:
        # Validate template exists
        template = db.query(models.ReportTemplate).filter(
            models.ReportTemplate.id == schedule_data.template_id
        ).first()
        if not template:
            raise HTTPException(status_code=404, detail="Report template not found")
        
        # Create new schedule
        schedule_dict = schedule_data.model_dump()
        schedule_dict["created_by"] = current_user.id
        schedule_dict["recipients"] = json.dumps(schedule_data.recipients)
        
        schedule = models.ScheduledDynamicReport(**schedule_dict)
        db.add(schedule)
        db.commit()
        db.refresh(schedule)
        
        # Convert response with parsed recipients
        try:
            recipients = json.loads(schedule.recipients) if schedule.recipients else []
        except json.JSONDecodeError:
            recipients = []
        
        # Get template data for response
        template_data = None
        if schedule.template:
            try:
                selected_columns = json.loads(schedule.template.selected_columns) if schedule.template.selected_columns else []
            except json.JSONDecodeError:
                selected_columns = []
            
            template_data = schemas.ReportTemplate(
                id=schedule.template.id,
                name=schedule.template.name,
                description=schedule.template.description,
                selected_columns=selected_columns,
                is_default=schedule.template.is_default,
                created_by=schedule.template.created_by,
                created_at=schedule.template.created_at,
                updated_at=schedule.template.updated_at
            )
        
        result = schemas.ScheduledDynamicReport(
            id=schedule.id,
            name=schedule.name,
            description=schedule.description,
            template_id=schedule.template_id,
            recipients=recipients,
            frequency=schedule.frequency,
            schedule_hour=schedule.schedule_hour,
            timezone=schedule.timezone,
            is_enabled=schedule.is_enabled,
            include_excel=schedule.include_excel,
            months_period=schedule.months_period,
            last_sent=schedule.last_sent,
            last_error=schedule.last_error,
            created_by=schedule.created_by,
            created_at=schedule.created_at,
            updated_at=schedule.updated_at,
            template=template_data
        )
        
        logger.info(f"Created scheduled dynamic report: {schedule.name}")
        return result
        
    except Exception as e:
        logger.error(f"Error creating scheduled dynamic report: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating scheduled report: {str(e)}")


@router.put("/schedules/{schedule_id}", response_model=schemas.ScheduledDynamicReport)
def update_scheduled_dynamic_report(
    schedule_id: int,
    schedule_data: schemas.ScheduledDynamicReportUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Update a scheduled dynamic report"""
    try:
        schedule = db.query(models.ScheduledDynamicReport).filter(
            models.ScheduledDynamicReport.id == schedule_id,
            models.ScheduledDynamicReport.created_by == current_user.id
        ).first()
        
        if not schedule:
            raise HTTPException(status_code=404, detail="Scheduled report not found")
        
        # Update fields
        update_dict = {k: v for k, v in schedule_data.model_dump().items() if v is not None}
        
        if "recipients" in update_dict:
            update_dict["recipients"] = json.dumps(update_dict["recipients"])
        
        for key, value in update_dict.items():
            setattr(schedule, key, value)
        
        schedule.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(schedule)
        
        logger.info(f"Updated scheduled dynamic report {schedule_id}: {schedule.name}")
        return schedule
        
    except Exception as e:
        logger.error(f"Error updating scheduled dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating scheduled report: {str(e)}")


@router.delete("/schedules/{schedule_id}")
def delete_scheduled_dynamic_report(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Delete a scheduled dynamic report"""
    try:
        schedule = db.query(models.ScheduledDynamicReport).filter(
            models.ScheduledDynamicReport.id == schedule_id,
            models.ScheduledDynamicReport.created_by == current_user.id
        ).first()
        
        if not schedule:
            raise HTTPException(status_code=404, detail="Scheduled report not found")
        
        db.delete(schedule)
        db.commit()
        
        logger.info(f"Deleted scheduled dynamic report {schedule_id}")
        return {"message": f"Scheduled report {schedule_id} deleted successfully"}
        
    except Exception as e:
        logger.error(f"Error deleting scheduled dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting scheduled report: {str(e)}")


@router.post("/schedules/{schedule_id}/toggle")
def toggle_scheduled_dynamic_report(
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    """Toggle enabled/disabled status of a scheduled dynamic report"""
    try:
        schedule = db.query(models.ScheduledDynamicReport).filter(
            models.ScheduledDynamicReport.id == schedule_id,
            models.ScheduledDynamicReport.created_by == current_user.id
        ).first()
        
        if not schedule:
            raise HTTPException(status_code=404, detail="Scheduled report not found")
        
        schedule.is_enabled = not schedule.is_enabled
        schedule.updated_at = datetime.now(timezone.utc)
        db.commit()
        
        status = "enabled" if schedule.is_enabled else "disabled"
        logger.info(f"Toggled scheduled dynamic report {schedule_id} to {status}")
        return {"message": f"Scheduled report {status} successfully", "is_enabled": schedule.is_enabled}
        
    except Exception as e:
        logger.error(f"Error toggling scheduled dynamic report: {e}")
        raise HTTPException(status_code=500, detail=f"Error toggling scheduled report: {str(e)}")
